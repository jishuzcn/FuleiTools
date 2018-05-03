#!/usr/bin/python
# -*- coding: utf-8 -*-
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import QDate
from ui import Ui_MainWindow, LoginDialog
import sys
from Apis import Spider
import LoginFulei
from Tools import FuleiTool
import time
import furl
from bs4 import BeautifulSoup
from openpyxl import Workbook


class MainUI(Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.spider = Spider()
        self.tools = FuleiTool()
        self.about_msg = "程序使用Python编写,用到的模块有PYQT5、Requests、BeautifulSoup4、re、OpenPyXL...\n\n" \
                         "本软件的设计初衷是为了解决个人工作中的某些任务，软件还有很多不足之处有待完善。\n\n" \
                         "程序分为三个功能，都涉及到利用爬虫技术爬取网站信息，因此一旦网页内容有较大变动，" \
                         "则原信息提取方式会失效，程序也会自动报错，如若程序无法继续使用，可以联系作者进行反馈！！\n\n" \
                         "编译日期：2018-5-1\n"
        self.author_msg = "作者：汪思源\n\n如果对本款小工具有疑问或者发现Bug，可以与本人联系交流！\n\n" \
                          "Github：https://github.com/jishuzcn\n\nBlog：http://code666.top\n\n联系QQ：11240532\n"

        self.actionQuit.triggered.connect(self.close)  # 菜单栏退出按钮函数
        self.actionAbout.triggered.connect(lambda: self.select_info("关于软件", self.about_msg))
        self.actionAuthor.triggered.connect(lambda: self.select_info("作者", self.author_msg))  # 关于作者
        self.textEdit_4.clicked.connect(self.clicked_textEdit)
        self.textEdit_5.clicked.connect(self.clicked_textEdit)
        self.pushButton_1.clicked.connect(self.start_run_more)
        self.textEdit_6.clicked.connect(self.clicked_textEdit)
        self.textEdit_7.clicked.connect(self.clicked_textEdit)
        self.pushButton_2.clicked.connect(self.start_run_one)

        dataHolder = self.spider.getAllHolderOrDealer()
        for key, value in dataHolder.items():
            self.comboBox.addItem(value, key)
        dataDealer = self.spider.getAllHolderOrDealer(HolderFlag=False)
        for key, value in dataDealer.items():
            self.comboBox_2.addItem(value, key)

    def closeEvent(self, event):
        reply = QtWidgets.QMessageBox.question(self, '关闭程序',
                                               "关闭程序可能导致正在进行的操作终止，请确认\n是否退出并关闭程序？",
                                               QtWidgets.QMessageBox.Yes, QtWidgets.QMessageBox.No)
        if reply == QtWidgets.QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

    def clicked_textEdit(self):
        self.textEditSender = self.sender()
        self.calendar = QtWidgets.QCalendarWidget(self.tabWidget)
        self.calendar.setGridVisible(True)
        self.calendar.show()
        date = self.calendar.selectedDate()
        self.calendar.clicked[QDate].connect(self.showDate)

    def showDate(self, date):
        self.textEditSender.setText(date.toString("yyyy-MM-dd"))
        if self.calendar.isHidden():
            self.calendar.setHidden(False)
        else:
            self.calendar.setHidden(True)

    def select_info(self, title, msg):
        QtWidgets.QMessageBox.about(self, title, msg)

    # 更新状态栏
    def statusshow(self, atr):
        self.statusbar.showMessage(atr)

    # 导出所有机器数据槽函数-------------------------------------------------------------------------------------------
    def start_run_more(self):
        self.statusbar.setStyleSheet("color:red")
        self.pushButton_1.setEnabled(False)
        self.textEdit_1.setText("")  # 插入一个空白,下次清屏
        self.progressBar_1.setValue(0)
        time1 = self.textEdit_4.toPlainText()
        time2 = self.textEdit_5.toPlainText()

        eid = self.textEdit.toPlainText()
        eName = self.textEdit_3.toPlainText()
        startTime = time1 if not time1 == "" else self.tools.timestampToTime(self.tools.getStartTimeOfToday())
        endTime = time2 if not time2 == "" else self.tools.timestampToTime(self.tools.getendTimeOfToday())
        combox1 = self.comboBox.itemData(self.comboBox.currentIndex())
        combox2 = self.comboBox_2.itemData(self.comboBox_2.currentIndex())
        holder = combox1 if not combox1 is None else ""
        dealer = combox2 if not combox2 is None else ""
        xlsName = self.lineEdit_1.text() + ".xlsx"
        self.morethread = runMoreThread(xlsName=xlsName, devicesId=eid, equipmentName=eName, startTime=startTime,
                                        endTime=endTime, holderId=holder, dealerId=dealer)
        # self.morethread = runMoreThread(startTime=startTime, endTime=endTime,dealerId=dealer,xlsName=xlsName)
        self.morethread.status_signal.connect(self.statusshow)
        self.morethread.text_signal.connect(self.moretextshow)
        self.morethread.progmax_signal.connect(self.moreprog_max)
        self.morethread.progcurrent_signal.connect(self.moreprog_value)
        self.morethread.finished.connect(self.morepushon)  # 线程结束执行函数
        self.morethread.start()

    # 更新输出文本
    def moretextshow(self, atr):
        self.textEdit_1.append(atr)

    # 进度条最大值
    def moreprog_max(self, n):
        self.progressBar_1.setMinimum(0)
        self.progressBar_1.setMaximum(n)

    # 进度条当前值
    def moreprog_value(self, i):
        self.progressBar_1.setValue(i)

    # 按钮变化
    def morepushon(self):
        self.pushButton_1.setDisabled(False)

    # 导出单个机器数据槽函数-------------------------------------------------------------------------------------------
    def start_run_one(self):
        self.statusbar.setStyleSheet("color:green")
        self.pushButton_2.setEnabled(False)
        self.textEdit_9.setText("")
        self.progressBar_2.setValue(0)
        id = self.textEdit_8.toPlainText()
        one_time1 = self.textEdit_6.toPlainText()
        one_time2 = self.textEdit_7.toPlainText()
        startTime = one_time1 if not one_time1 == "" else self.tools.timestampToTime(self.tools.getStartTimeOfToday())
        endTime = one_time2 if not one_time2 == "" else self.tools.timestampToTime(self.tools.getEndTimeOfToday())
        xlsName = self.lineEdit_2.text() + ".xlsx"
        if id == "":
            self.select_info("Warning!!!", "你的id为空，这是不合法的")
            self.pushButton_2.setEnabled(True)
        else:
            self.onethread = runOneThread(xlsName=xlsName,id=id,startTime=startTime,endTime=endTime)
            self.onethread.status_signal.connect(self.statusshow)
            self.onethread.text_signal.connect(self.onetextshow)
            self.onethread.progmax_signal.connect(self.oneprog_max)
            self.onethread.progcurrent_signal.connect(self.oneprog_value)
            self.onethread.finished.connect(self.onepushon)  # 线程结束执行函数
            self.onethread.start()

    # 更新输出文本
    def onetextshow(self, atr):
        self.textEdit_9.append(atr)

    # 进度条最大值
    def oneprog_max(self, n):
        self.progressBar_2.setMinimum(0)
        self.progressBar_2.setMaximum(n)

    # 进度条当前值
    def oneprog_value(self, i):
        self.progressBar_2.setValue(i)

    # 按钮变化
    def onepushon(self):
        self.pushButton_2.setDisabled(False)

class runOneThread(QtCore.QThread):
    status_signal = QtCore.pyqtSignal(str)  # 发送给状态栏信号
    text_signal = QtCore.pyqtSignal(str)  # 发送给输入框信号
    progmax_signal = QtCore.pyqtSignal(int)  # 发送给进度条信号，给出最大值
    progcurrent_signal = QtCore.pyqtSignal(int)  # 发送给进度条信号，给出当前进度

    def __init__(self,xlsName="bbb.xlsx",**kwargs):
        super().__init__()
        self.spider = Spider()
        self.xlsName = xlsName
        self.url = "http://www.kadawo.com/fulei/index.php/equipment/outPut/page/eType1information/szType/zx/did/{}/" \
                   "startTime/{}/endTime/{}".format(kwargs['id'], kwargs['startTime'], kwargs['endTime'])

    def run(self):
        "equipment/outPut/page/eType1information/szType/zx/did/867793034475113/startTime/2018-03-01/endTime/2018-04-01"
        header = {'Host': 'www.kadawo.com','Proxy-Connection': 'keep-alive',
                  'Upgrade-Insecure-Requests': '1',
                  'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36',
                  'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
                  'Accept-Encoding': 'gzip, deflate',
                  'Accept-Language': 'zh-CN,zh;q=0.9'
                  }
        header.setdefault('Cookie', 'PHPSESSID=%s' % self.spider.login.getCookie())
        start = time.time()
        self.status_signal.emit("当前状态:开始写入数据...")
        self.progmax_signal.emit(2)
        self.progcurrent_signal.emit(1)
        self.text_signal.emit("速度很快，请等待...")
        f = open(self.xlsName, 'wb')
        f.write(self.spider.login.session.get(self.url, headers=header, verify=False).content)
        f.close()
        end = time.time()
        msg_d = "操作完毕，耗时：%0.2f秒！\n数据保存在当前目录下文件  %s 中" % (float(end - start), self.xlsName)
        msg_e = "如果文件错误，请复制下方链接到浏览器下载\n%s" % self.url
        self.progcurrent_signal.emit(2)
        self.text_signal.emit(self.spider.getmsg(msg_d, "green"))
        self.text_signal.emit(self.spider.getmsg(msg_e, "red"))
        self.status_signal.emit("当前状态:写入完毕...")


class runMoreThread(QtCore.QThread):
    status_signal = QtCore.pyqtSignal(str)  # 发送给状态栏信号
    text_signal = QtCore.pyqtSignal(str)  # 发送给输入框信号
    progmax_signal = QtCore.pyqtSignal(int)  # 发送给进度条信号，给出最大值
    progcurrent_signal = QtCore.pyqtSignal(int)  # 发送给进度条信号，给出当前进度

    def __init__(self, xlsName="aaa.xlsx", **kwargs):
        super().__init__()
        self.spider = Spider()
        self.tools = FuleiTool()
        self.wb = Workbook()
        kwargs_key = list(kwargs.keys())
        for kk in kwargs_key:
            if kwargs[kk] == "":
                del kwargs[kk]
        f = furl.furl('http://www.kadawo.com/fulei/index.php/equipment/shjList')
        f.args = kwargs
        f.args.setdefault('__hash__', self.tools.getHash(self.spider.getHtml()))
        self.url = self.tools.urlReplace(f.url)
        print(self.url)
        self.startTime = kwargs['startTime']
        self.endTime = kwargs['endTime']
        self.xlsName = xlsName

    def run(self):
        start = time.time()
        self.status_signal.emit("当前状态:正在进行爬取写入操作...")

        ws = self.wb.active
        ws.title = "More Sheet"
        ws.sheet_properties.tabColor = "1072BA"
        ws['A1'] = "导出时间(%s,%s)" % (self.startTime, self.endTime)
        ws['A2'] = "设备id"
        ws['B2'] = "设备名称"
        ws['C2'] = "设备地址"
        ws['D2'] = "持有人"
        ws['E2'] = "所属"
        ws['F2'] = "合计次数"
        ws['G2'] = "合计金额"
        ws['H2'] = "库存"
        ws['I2'] = "运行状态"
        ws['J2'] = "当前彩票品种"

        startT = self.tools.get_startTime(self.tools.timeToTimestamp(self.startTime))
        endT = self.tools.get_endTime(self.tools.timeToTimestamp(self.endTime))
        htm = self.spider.getHtml(self.url)
        data = self.writeData(htm, startT, endT)
        totalPage = int(self.tools.getTotalPage(htm))
        n = 1
        self.progmax_signal.emit(totalPage)
        self.progcurrent_signal.emit(n)
        self.text_signal.emit("总计 {} 页数据,成功写入第 {} 页数据".format(totalPage, n))
        while totalPage > n:
            n += 1
            self.progcurrent_signal.emit(n)
            next_url = self.url + '/p/' + str(n) + '/'
            htm = self.spider.getHtml(next_url)
            self.writeData(htm, startT, endT)  # error!!! ChunkedEncodingError
            self.text_signal.emit("总计 {} 页数据,成功写入第 {} 页数据".format(totalPage, n))
        self.wb.save(self.xlsName)
        end = time.time()
        msg_d = "操作完毕，耗时：%0.2f秒！\n数据保存在当前目录下文件  %s 中" % (float(end - start), self.xlsName)
        self.text_signal.emit(self.spider.getmsg(msg_d, "green"))
        self.status_signal.emit("当前状态:操作完毕...")

    def writeData(self, htm, startTime=None, endTime=None):
        ws = self.wb.active
        if not startTime:
            startTime = self.login.fltool.getStartTimeOfToday()
        if not endTime:
            endTime = self.login.fltool.getEndTimeOfToday()
        soup = BeautifulSoup(htm, "lxml")
        tbody = soup.find("tbody")
        rows = tbody.findAll('tr')
        # 36 0441 100280036 055
        # 09704933021
        for row in rows:
            cols = row.findAll('td')
            if len(cols) > 15:
                da = []
                da.append(cols[2].text.strip())
                da.append(cols[3].span.attrs['title'] if not cols[3].span.attrs['title'] == '' else '')
                da.append(cols[4].span.attrs['title'] if not cols[4].span.attrs['title'] == '' else '')
                da.append(cols[7].text.strip())
                da.append(cols[8].text.strip())
                ttmoney = self.tools.getMoney(cols[2].text.strip(), startTime, endTime)
                da.append(ttmoney[0])
                da.append(ttmoney[1])
                da.append(cols[11].text.strip())
                da.append(cols[12].text.strip())
                da.append(self.spider.get_lottery_name(cols[1].text.strip()))
                ws.append(da)


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    login = LoginFulei.LoginFulei()
    dialog = LoginDialog()
    if login.isLogin():
        my = MainUI()
        my.show()
        sys.exit(app.exec_())
    else:
        if dialog.exec_():
            my = MainUI()
            my.show()
            sys.exit(app.exec_())
